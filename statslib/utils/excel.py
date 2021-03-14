# http://docs.xlwings.org/en/0.13.0/datastructures.html
# http://docs.xlwings.org/en/stable/api.html?#xlwings.main.Sheets.add

import pandas as pd
import os
from statslib.utils.file import create_file_name
from statslib.utils.log import Logger
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

def _error_to_float(row):
    res = []
    for r in row:
        try:
            res.append(float(r))
        except:
            res.append(r)
    return res


def export_obj_to_excel(obj, book=None, sheet='Sheet1', **kwargs):
    if book is None:
        book = create_file_name(extension='xlsx', silent=True)
    if os.path.exists(book):
        wb = load_workbook(book)
    else:
        wb = Workbook()
    ws_active = wb.active
    if ws_active.title == 'Sheet': wb.remove(ws_active)
    ws = wb.create_sheet(sheet)

    if isinstance(obj, pd.DataFrame) or isinstance(obj, pd.Series):
        obj.columns = [str(col) for col in obj.columns]
        for row in dataframe_to_rows(obj, index=False, header=True):
            row = [str(r) if row is not None else row for r in row]
            row = _error_to_float(row)
            ws.append(row)
        wb.save(book)
    elif isinstance(obj, list):
        export_obj_to_excel(pd.DataFrame(obj), book, sheet, **kwargs)
    else:
        try:
            ws['A1'] = obj
        except Exception as e:
            try:
                ws['A1'] = obj.__str__()
            except Exception as e:
                err_str = 'no error message.'
                Logger().get_logger().debug("Can't export object!")
                try:
                    err_str = e.args[0]
                except:
                    pass
                ws['A1'] = "Cant import object: " + err_str
    wb.save(book)


# def export_obj_to_excel(df, book=None, sheet='Sheet1', **kwargs):
#     if isinstance(df, pd.DataFrame) or isinstance(df, pd.Series):
#         if book is None:
#             book = create_file_name(extension='xlsx', silent=False)
#
#         workbook = xlsxwriter.Workbook(book)
#         worksheet = workbook.add_worksheet(sheet)
#
#         writer = pd.ExcelWriter(book, engine='xlsxwriter')
#         df.to_excel(writer, sheet_name=sheet, **kwargs)
#         writer.save()

# @xw.ret(header=True, expand='table')
# def export_obj_to_excel(obj,
#                         book=None,
#                         sheet='Sheet1',
#                         range='A1',
#                         quit_after_done=False,
#                         after_sheet=None,
#                         return_objs=False):
#     if book is None:
#         book = create_file_name(extension='xlsx', silent=True)
#
#     if not os.path.exists(book):
#         n_book = xw.Book()
#         n_book.save(book)
#         n_book.close()
#
#     m_book = xw.Book(book)
#     m_app = xw.apps.active
#     m_sheets = [val.name for val in m_book.sheets]
#     if sheet in m_sheets: m_book.sheets[sheet].delete()
#     if after_sheet not in m_sheets:
#         Logger().get_logger().debug("after_sheet {} is not in Workbook sheets".format(after_sheet))
#         after_sheet = None
#     m_book.sheets.add(sheet, after=m_sheets[-1] if after_sheet is None else after_sheet)
#     m_sheets = [val.name for val in m_book.sheets]
#     if 'Sheet1' in m_sheets: m_book.sheets['Sheet1'].delete()
#
#     if isinstance(obj, pd.DataFrame) or isinstance(obj, pd.Series):
#         _ = m_book.sheets[sheet].range(range).options(
#             index=False if obj.index.name is None else True).value = obj
#     else:
#         if isinstance(obj, list):
#             obj = pd.DataFrame(obj)
#         else:
#             try:
#                 obj = pd.DataFrame(obj)
#             except:
#                 pass
#         _ = m_book.sheets[sheet].range(range).value = obj
#
#     m_book.save(book)
#
#     if quit_after_done: m_app.quit()
#
#     if return_objs: return m_app, book, m_sheets
